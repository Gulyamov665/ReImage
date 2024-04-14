import os
import zipfile
from PIL import Image
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
import openpyxl
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.generics import CreateAPIView, ListAPIView
from rest_framework.views import APIView
from django.conf import settings
from multi_upload.models import Images, Restaurant, PromoSticker, VendorImage
from multi_upload.serializers import ProductSerializers, PromoSerializers, Promo
from django.contrib.auth.models import User
from rest_framework.viewsets import ModelViewSet, ReadOnlyModelViewSet
from multi_upload import models
from multi_upload import serializers
from rest_framework.parsers import (
    JSONParser,
    FormParser,
    MultiPartParser,
)


# class ProductCreateView(CreateAPIView):
#     parser_class = [MultiPartParser, FormParser]
#     serializer_class = ProductSerializers


# class ImageResizeView(APIView):
#     def post(self, request):
#         result = request.data.get("res")
#         res_width = 1500
#         input_folder = os.path.join(settings.MEDIA_ROOT, str(result))
#         output_folder = os.path.join(settings.MEDIA_ROOT, "output", str(result))
#         if not os.path.exists(output_folder):
#             os.makedirs(output_folder)
#         name = 0

#         for file in os.listdir(input_folder):
#             name += 1
#             filename = str(name)
#             files = (".jpg", ".png", ".jpeg", ".webp", ".jfif")
#             if file.endswith(files):
#                 img = Image.open(os.path.join(input_folder, file))
#                 wpercent = res_width / float(img.size[0])
#                 hsize = int((float(img.size[1]) * float(wpercent)))
#                 img = img.resize((res_width, hsize), Image.Resampling.LANCZOS)
#                 try:
#                     img.save(os.path.join(output_folder, filename + ".jpg"))
#                 except OSError:
#                     img.save(os.path.join(output_folder, filename + ".png"))

#         return Response({"message": "Images resized successfully"})


# class ImagesApiView(ModelViewSet):
#     queryset = Images.objects.all()
#     parser_class = [MultiPartParser, FormParser]
#     serializer_class = ProductSerializers


# class DetailUserApiView(ListAPIView):
#     parser_class = [MultiPartParser, FormParser]
#     serializer_class = ProductSerializers

#     def get_queryset(self):
#         user = self.kwargs["user_id"]
#         queryset = Restaurant.objects.filter(user_id=user)
#         return queryset


# class DetailAllApiView(ListAPIView):
#     parser_class = [MultiPartParser, FormParser]
#     queryset = Restaurant.objects.all()
#     serializer_class = ProductSerializers


# class DownloadFilesView(APIView):
#     def get(self, request, name):
#         response = HttpResponse(content_type="application/zip")
#         response["Content-Disposition"] = "attachment; filename=files.zip"

#         with zipfile.ZipFile(response, "w") as zf:
#             for filename in os.listdir(os.path.join("media/output", name)):
#                 zf.write(os.path.join("media/output", name, filename), filename)

#         return response


# def export_xlsx(request, name):
#     rest = get_object_or_404(Restaurant, name=name)
#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = 'attachment; filename="vendor.xlsx"'

#     wb = openpyxl.Workbook()
#     ws = wb.active

#     ws.cell(row=1, column=1, value="Image")
#     ws.cell(row=1, column=2, value="Name")
#     ws.cell(row=1, column=3, value="Description")
#     ws.cell(row=1, column=4, value="Price")

#     number = os.path.join(settings.MEDIA_ROOT, "output", name)
#     num_name = os.listdir(number)
#     photo_name = Images.objects.filter(product=rest).values_list("img_title", flat=True)
#     price = Images.objects.filter(product=rest).values_list("img_price", flat=True)
#     desc = Images.objects.filter(product=rest).values_list("img_description", flat=True)

#     for d, num_name in enumerate(num_name):
#         row = d + 2
#         ws.cell(row=row, column=1, value=num_name)

#     for i, photo_name in enumerate(photo_name):
#         row = i + 2
#         ws.cell(row=row, column=2, value=photo_name)

#     for e, desc in enumerate(desc):
#         row = e + 2
#         ws.cell(row=row, column=3, value=desc)

#     for c, price in enumerate(price):
#         row = c + 2
#         ws.cell(row=row, column=4, value=price)

#     wb.save(response)
#     return response


# @api_view(["GET"])
# def apiOverview(request):
#     api_urls = {
#         "Create": "api/create",
#         "Detail/user": "api/detail/<int:user_id>",
#         "Detail/all": "api/detail",
#         "Create User": "api/create_user",
#         "Resize": "api/resize_image",
#     }
#     return Response(api_urls)


# class PromoStickerView(ModelViewSet):
#     queryset = PromoSticker.objects.all()
#     serializer_class = PromoSerializers
#     parser_classes = [JSONParser, FormParser, MultiPartParser]

#     def sticker(self, request):
#         name_path = request.data.get("vendor")
#         promo_link = request.data.get("tag")

#         print(promo_link)

#         promo_folder = os.path.join(settings.MEDIA_ROOT, f"stickers")
#         origin_folder = os.path.join(settings.MEDIA_ROOT, f"{name_path}/origin")
#         output_folder = os.path.join(settings.MEDIA_ROOT, f"{name_path}/with_tag")
#         os.makedirs(output_folder, exist_ok=True)
#         os.makedirs(origin_folder, exist_ok=True)

#         for image_name in os.listdir(origin_folder):
#             original_image = Image.open(f"{origin_folder}/{image_name}")
#             try:
#                 original_image = original_image.resize((1800, 1080))
#                 promo_files = os.listdir(promo_folder)
#                 if promo_link:
#                     first_file = promo_files[0]
#                 tag_image = Image.open(promo_link)
#                 if tag_image.mode == "RGBA":
#                     original_image.paste(tag_image, (0, 0), tag_image)

#                     original_image.save(os.path.join(output_folder, image_name))
#                     print(f"Обработано: {image_name}")
#                     # удаление папок или файлов
#                     # os.remove(first_file)
#                     # os.rmdir(origin_folder)
#                 else:
#                     print(f"Изображение тега не имеет прозрачного фона: {image_name}")
#             except Exception as e:
#                 print(f"Ошибка при обработке {image_name}: {str(e)}")
#         return Response(
#             {"message": "Все фотографии обработаны и сохранены в папке with_tag_new."}
#         )

#     def download(self, request, name):
#         response = HttpResponse(content_type="application/zip")
#         response["Content-Disposition"] = "attachment; filename=files.zip"

#         with zipfile.ZipFile(response, "w") as zf:
#             dir_path = os.path.join("media", name, "with_tag")
#             for root, _, files in os.walk(dir_path):
#                 for filename in files:
#                     file_path = os.path.join(root, filename)
#                     zf.write(file_path, arcname=os.path.relpath(file_path, dir_path))

#         return response


# class VendorImageViews(ModelViewSet):
#     queryset = VendorImage.objects.all()
#     serializer_class = PromoSerializers
#     parser_classes = [JSONParser]


# class PromoView(ReadOnlyModelViewSet):
#     queryset = PromoSticker.objects.all()
#     serializer_class = Promo


from urllib.request import urlopen


class TagsView(ReadOnlyModelViewSet):
    queryset = models.Tag.objects.all()
    serializer_class = serializers.TagsSerializer


class VendorView(ModelViewSet):
    queryset = models.Vendor.objects.all()
    serializer_class = serializers.VendorSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def sticker(self, request):
        name_path = request.data.get("vendor")
        promo_link = request.data.get("tag")

        print(promo_link)

        # promo_folder = os.path.join(settings.MEDIA_ROOT, f"stickers")
        origin_folder = os.path.join(settings.MEDIA_ROOT, f"{name_path}/origin")
        output_folder = os.path.join(settings.MEDIA_ROOT, f"{name_path}/with_tag")
        os.makedirs(output_folder, exist_ok=True)
        os.makedirs(origin_folder, exist_ok=True)

        for image_name in os.listdir(origin_folder):
            original_image = Image.open(f"{origin_folder}/{image_name}")
            try:
                original_image = original_image.resize((1800, 1080))
                # promo_files = os.listdir(promo_folder)
                # if promo_link:
                #     first_file = promo_files[0]
                tag_image = Image.open(urlopen(promo_link))
                if tag_image.mode == "RGBA":
                    original_image.paste(tag_image, (0, 0), tag_image)

                    original_image.save(os.path.join(output_folder, image_name))
                    print(f"Обработано: {image_name}")
                    # удаление папок или файлов
                    # os.remove(first_file)
                    # os.rmdir(origin_folder)
                else:
                    print(f"Изображение тега не имеет прозрачного фона: {image_name}")
            except Exception as e:
                print(f"Ошибка при обработке {image_name}: {str(e)}")
        return Response(
            {"message": "Все фотографии обработаны и сохранены в папке with_tag_new."}
        )

    def download(self, request, name):
        response = HttpResponse(content_type="application/zip")
        response["Content-Disposition"] = "attachment; filename=files.zip"

        with zipfile.ZipFile(response, "w") as zf:
            dir_path = os.path.join("media", name, "with_tag")
            for root, _, files in os.walk(dir_path):
                for filename in files:
                    file_path = os.path.join(root, filename)
                    zf.write(file_path, arcname=os.path.relpath(file_path, dir_path))

        return response
